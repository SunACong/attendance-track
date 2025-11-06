import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox
import shutil
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import Workbook

from all import build_record_index, init_attendance_template, summarize_attendance
from processCCKQ import fill_business_trip
from processLGDJ import fill_leave_registration
from processPCKQ import fill_pc_attendance, process_pc_attendance
from processQJDJ import fill_leave_info
from processShift import fill_shift_attendance
from processYDKQ import fill_oa_attendance

files = {}
labels = {}
status_label = None

# === 在保存汇总表之前，清理 0 ===
def clean_zeros(df):
    return df.applymap(lambda x: "" if (isinstance(x, (int, float)) and x == 0) else x)

# === 保存带颜色标记的Excel文件 ===
def save_excel_with_highlight(df, file_path):
    # 创建ExcelWriter对象
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    # 将DataFrame写入Excel
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    # 获取工作表对象
    worksheet = writer.sheets['Sheet1']
    
    # 查找'是否异常'列的索引
    abnormal_col = None
    for col_idx, col_name in enumerate(df.columns):
        if col_name == '是否异常':
            abnormal_col = col_idx + 1  # openpyxl列索引从1开始
            break
    
    # 如果找到'是否异常'列，添加颜色标记
    if abnormal_col is not None:
        # 创建填充样式（黄色背景）
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # 遍历所有行，标记'是否异常'为'是'的行
        for row_idx in range(2, len(df) + 2):  # 从第二行开始（第一行是表头）
            cell = worksheet.cell(row=row_idx, column=abnormal_col)
            if cell.value == '是':
                # 标记整行
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col).fill = fill
    
    # 保存文件
    writer.close()

def upload_file(key):
    path = filedialog.askopenfilename(filetypes=[("Excel or CSV files", "*.xlsx *.csv")])
    if path:
        files[key] = path
        labels[key].config(text=os.path.basename(path))



def update_status(root, message):
    status_label.config(text=message)
    root.update_idletasks()



def run_analysis(root):
    try:
        required_keys = ["person", "oa", "trip", "pc", "leave", "shift", "qj", "holiday", "record"]
        if not all(k in files for k in required_keys):
            messagebox.showerror("缺少文件", "请确保已选择所有所需文件。")
            return

        start_time = time.time()
        update_status(root, "🕐 正在加载数据...")

        person_df = pd.read_excel(files["person"], dtype={"工号": str})
        oa_df = pd.read_excel(files["oa"], dtype={"编号": str})
        leave_df = pd.read_excel(files["leave"], dtype={"人员编码": str})
        qj_df = pd.read_excel(files["qj"], dtype={"工号": str})
        holiday_df = pd.read_excel(files["holiday"])
        holiday_set = set(pd.to_datetime(holiday_df["日期"]).dt.date)
        trip_df = pd.read_excel(files["trip"], dtype={"人员编号": str})

        if files["shift"].endswith(".xlsx"):
            shift_df = pd.read_excel(files["shift"], dtype={"工号": str})
        else:
            shift_df = pd.read_csv(files["shift"], encoding="gbk", dtype={"工号": str})

        if files["record"].endswith(".csv"):
            record_df = pd.read_csv(files["record"], encoding="gbk", parse_dates=["考勤时间"], dtype={"工号": str})
        else:
            record_df = pd.read_excel(files["record"], dtype={"工号": str})

        update_status(root, "📊 正在处理 PC 考勤结果...")
        date_range, attendance_data = process_pc_attendance(files["pc"])
        contact_attendance_list, person_dept_dict = init_attendance_template(person_df, date_range[0], date_range[1])
        index_map = build_record_index(contact_attendance_list)

        fill_pc_attendance(index_map, attendance_data)
        update_status(root, "📊 正在处理 OA 考勤...")
        fill_oa_attendance(index_map, oa_df)

        update_status(root, "📊 正在处理离岗登记...")
        fill_leave_registration(index_map, leave_df)
        update_status(root, "📊 正在处理请假记录...")
        print("📊 正在处理请假记录...")
        fill_leave_info(index_map, qj_df)
        update_status(root, "📊 正在处理出差记录...")
        fill_business_trip(index_map, trip_df)
        update_status(root, "📊 正在处理倒班记录...")
        shift_day_dict = fill_shift_attendance(index_map, shift_df, record_df, holiday_set, person_dept_dict)

        update_status(root, "📊 正在汇总数据...")
        summary_result = summarize_attendance(contact_attendance_list, holiday_set, shift_day_dict)
        df_summary = pd.DataFrame(summary_result)
        # 打印holiday_set前五条数据
        # print(holiday_set[:5])
        df_all = pd.DataFrame(contact_attendance_list)

        update_status(root, "💾 正在保存结果...")
        save_base = filedialog.asksaveasfilename(title="保存结果文件", defaultextension=".xlsx",
                                                  filetypes=[("Excel 文件", "*.xlsx")])
        if save_base:
            # 根目录
            base_dir = os.path.dirname(save_base)

            # === 清空根目录下的旧文件/文件夹 ===
            for item in os.listdir(base_dir):
                item_path = os.path.join(base_dir, item)
                try:
                    if os.path.isfile(item_path) or os.path.islink(item_path):
                        os.remove(item_path)
                    elif os.path.isdir(item_path):
                        shutil.rmtree(item_path)
                except Exception as e:
                    print(f"清理失败: {item_path}, {e}")

            # === 保存新的结果 ===
            summary_path = os.path.join(base_dir, "所有单位汇总表.xlsx")
            detail_path = os.path.join(base_dir, "所有单位明细表.xlsx")
            df_summary = clean_zeros(df_summary)  # 汇总表清理
            
            # 使用带颜色标记的保存函数
            update_status(root, "💾 正在保存带颜色标记的汇总表...")
            save_excel_with_highlight(df_summary, summary_path)
            
            update_status(root, "💾 正在保存带颜色标记的明细表...")
            save_excel_with_highlight(df_all, detail_path)

            # 创建子目录
            summary_dir = os.path.join(base_dir, "各单位汇总表")
            detail_dir = os.path.join(base_dir, "各单位明细表")
            os.makedirs(summary_dir, exist_ok=True)
            os.makedirs(detail_dir, exist_ok=True)

            # 按一级部门分组并导出
            if "部门" in df_summary.columns:
                dept_groups_summary = df_summary.groupby(df_summary["部门"].astype(str).str.split("/").str[0])
                dept_groups_detail = df_all.groupby(df_all["部门"].astype(str).str.split("/").str[0])

                for dept, group in dept_groups_summary:
                    dept_name = str(dept).strip().replace("/", "_").replace("\\", "_")
                    dept_file = os.path.join(summary_dir, f"{dept_name}_汇总.xlsx")
                    save_excel_with_highlight(group, dept_file)

                for dept, group in dept_groups_detail:
                    dept_name = str(dept).strip().replace("/", "_").replace("\\", "_")
                    dept_file = os.path.join(detail_dir, f"{dept_name}_明细.xlsx")
                    save_excel_with_highlight(group, dept_file)

                update_status(
                    root,
                    f"✅ 已拆分完成，共 {df_summary['部门'].astype(str).str.split('/').str[0].nunique()} 个一级部门"
                )

        elapsed = time.time() - start_time
        update_status(root, f"✅ 分析完成，用时 {elapsed:.2f} 秒。")

    except Exception as e:
        messagebox.showerror("出错啦", str(e))
        update_status(root, "❌ 分析失败")



def main():
    global status_label

    root = tk.Tk()
    root.title("📊 考勤分析工具 (Tkinter 版)")

    frame = tk.Frame(root, padx=12, pady=12)
    frame.pack()

    file_keys = [
        ("person", "通信录"), ("oa", "OA打卡"), ("trip", "出差记录"),
        ("pc", "PC考勤结果"), ("leave", "离岗登记"), ("shift", "倒班记录"),
        ("qj", "请假记录"), ("holiday", "节假日"), ("record", "PC打卡记录")
    ]

    for key, name in file_keys:
        row = tk.Frame(frame)
        row.pack(fill="x", pady=2)
        tk.Label(row, text=name, width=15, anchor="w").pack(side="left")
        labels[key] = tk.Label(row, text="未选择文件", width=40, anchor="w", relief="sunken")
        labels[key].pack(side="left")
        tk.Button(row, text="选择", command=lambda k=key: upload_file(k)).pack(side="left")

    tk.Button(frame, text="🚀 开始分析", bg="#28a745", fg="white",
              command=lambda: run_analysis(root)).pack(pady=10)

    status_label = tk.Label(root, text="等待操作...", fg="blue", anchor="w")
    status_label.pack(fill="x", padx=12, pady=(0, 12))

    root.mainloop()


if __name__ == "__main__":
    main()
