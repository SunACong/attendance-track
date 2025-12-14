import os
import time
import zipfile
import streamlit as st
import pandas as pd
import tempfile
from openpyxl.styles import PatternFill
import shutil

# 导入现有的处理函数
from all import build_record_index, init_attendance_template, summarize_attendance
from processCCKQ import fill_business_trip
from processLGDJ import fill_leave_registration
from processPCKQ import fill_pc_attendance, process_pc_attendance
from processQJDJ import fill_leave_info
from processShift import fill_shift_attendance
from processYDKQ import fill_oa_attendance

# 设置页面配置
st.set_page_config(
    page_title="考勤分析工具",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 页面标题
st.title("📊 考勤分析工具")

# 文件类型映射：关键字 -> 文件描述
FILE_TYPE_MAPPING = {
    "通信录": "person",
    "OA打卡": "oa",
    "出差记录": "trip",
    "PC考勤结果": "pc",
    "离岗登记": "leave",
    "倒班记录": "shift",
    "请假记录": "qj",
    "节假日": "holiday",
    "PC打卡记录": "record"
}

# === 在保存汇总表之前，清理 0 ===
@st.cache_data
def clean_zeros(df):
    return df.applymap(lambda x: "" if (isinstance(x, (int, float)) and x == 0) else x)

# === 保存带颜色标记的Excel文件 ===
def save_excel_with_highlight(df, file_path):
    # 创建ExcelWriter对象
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    # 将DataFrame写入Excel
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    # 获取工作表对象
    worksheet = writer.sheets['Sheet1']
    
    # 查找'是否异常'列的索引
    abnormal_col = None
    for col_idx, col_name in enumerate(df.columns):
        if col_name == '是否异常':
            abnormal_col = col_idx + 1  # openpyxl列索引从1开始
            break
    
    # 如果找到'是否异常'列，添加颜色标记
    if abnormal_col is not None:
        # 创建填充样式（黄色背景）
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # 遍历所有行，标记'是否异常'为'是'的行
        for row_idx in range(2, len(df) + 2):  # 从第二行开始（第一行是表头）
            cell = worksheet.cell(row=row_idx, column=abnormal_col)
            if cell.value == '是':
                # 标记整行
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col).fill = fill
    
    # 保存文件
    writer.close()

# === 拆分原始打卡记录 ===
def split_attendance_records(input_file, output_dir):
    """
    按二级组织拆分考勤记录文件
    :param input_file: 输入的考勤记录文件（CSV或Excel）
    :param output_dir: 拆分后文件的存储目录
    :return: 拆分后的文件列表
    """
    # 读取文件
    if input_file.name.endswith('.csv'):
        df = pd.read_csv(input_file, encoding='gbk')
    else:
        df = pd.read_excel(input_file)
    
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 从所属组织列中提取二级组织
    df['二级组织'] = df['所属组织'].str.split('/').str[1]
    
    # 按二级组织分组并保存文件
    split_files = []
    grouped = df.groupby('二级组织')
    
    for org_name, group in grouped:
        # 移除临时添加的二级组织列
        group = group.drop(columns=['二级组织'])
        output_file_path = os.path.join(output_dir, f'{org_name}_考勤记录.csv')
        group.to_csv(output_file_path, index=False, encoding='utf-8-sig')
        split_files.append(output_file_path)
    
    return split_files

# === 创建ZIP文件 ===
def create_zip_file(zip_filename, summary_file, detail_file, dept_summary_files, dept_detail_files, split_files=[]):
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        # 添加整体汇总表和明细表
        if os.path.exists(summary_file):
            zipf.write(summary_file, os.path.basename(summary_file))
        if os.path.exists(detail_file):
            zipf.write(detail_file, os.path.basename(detail_file))
        
        # 添加各单位汇总文件夹和文件
        for dept_name, file_path in dept_summary_files:
            if os.path.exists(file_path):
                zipf.write(file_path, f"各单位汇总/{dept_name}_汇总.xlsx")
        
        # 添加各单位明细文件夹和文件
        for dept_name, file_path in dept_detail_files:
            if os.path.exists(file_path):
                zipf.write(file_path, f"各单位明细/{dept_name}_明细.xlsx")
        
        # 添加原始打卡记录文件夹和拆分后的文件
        for file_path in split_files:
            if os.path.exists(file_path):
                file_name = os.path.basename(file_path)
                zipf.write(file_path, f"原始打卡记录/{file_name}")

# === 批量文件上传处理 ===
def process_uploaded_files(uploaded_files):
    files = {}
    unmatched_files = []
    
    for file in uploaded_files:
        file_name = file.name
        matched = False
        
        for keyword, key in FILE_TYPE_MAPPING.items():
            if keyword in file_name:
                files[key] = file
                matched = True
                break
        
        if not matched:
            unmatched_files.append(file_name)
    
    return files, unmatched_files

# 主界面布局
col1, col2 = st.columns([2, 1])

with col1:
    st.header("📁 文件上传")
    st.write("请上传所有考勤相关文件，系统会自动识别文件类型")
    
    # 批量文件上传控件
    uploaded_files = st.file_uploader(
        "选择文件",
        type=["xlsx", "csv"],
        accept_multiple_files=True,
        help="支持同时上传多个Excel或CSV文件"
    )

with col2:
    st.header("📋 文件类型说明")
    st.write("请在文件名中包含以下关键字：")
    for keyword, description in FILE_TYPE_MAPPING.items():
        st.write(f"- **{keyword}**: {description}")

# 处理文件上传
if uploaded_files:
    st.subheader("📊 文件识别结果")
    
    files, unmatched_files = process_uploaded_files(uploaded_files)
    
    # 显示匹配的文件
    if files:
        st.success(f"✅ 成功识别 {len(files)} 个文件")
        for key, file in files.items():
            st.write(f"- **{key}**: {file.name}")
    
    # 显示未匹配的文件
    if unmatched_files:
        st.warning(f"⚠️ 无法识别 {len(unmatched_files)} 个文件")
        for file_name in unmatched_files:
            st.write(f"- {file_name}")
    
    # 检查是否所有必需的文件都已上传
    required_keys = ["person", "oa", "trip", "pc", "leave", "shift", "qj", "holiday", "record"]
    missing_keys = [key for key in required_keys if key not in files]
    
    if missing_keys:
        st.error(f"❌ 缺少以下必需文件：{', '.join(missing_keys)}")
    else:
        st.success("✅ 所有必需文件已上传完成")
        
        # 初始化会话状态
        if 'analysis_completed' not in st.session_state:
            st.session_state.analysis_completed = False
        if 'df_summary' not in st.session_state:
            st.session_state.df_summary = None
        if 'df_all' not in st.session_state:
            st.session_state.df_all = None
        if 'dept_summary_files' not in st.session_state:
            st.session_state.dept_summary_files = []
        if 'dept_detail_files' not in st.session_state:
            st.session_state.dept_detail_files = []
        if 'zip_file_created' not in st.session_state:
            st.session_state.zip_file_created = False
        
        # 开始分析按钮
        if st.button("🚀 开始分析", key="start_analysis", help="点击开始处理考勤数据"):
            with st.spinner("🕐 正在处理考勤数据..."):
                try:
                    start_time = time.time()
                    
                    # 加载数据
                    person_df = pd.read_excel(files["person"], dtype={"工号": str})
                    oa_df = pd.read_excel(files["oa"], dtype={"编号": str})
                    leave_df = pd.read_excel(files["leave"], dtype={"人员编码": str})
                    qj_df = pd.read_excel(files["qj"], dtype={"工号": str})
                    holiday_df = pd.read_excel(files["holiday"])
                    holiday_set = set(pd.to_datetime(holiday_df["日期"]).dt.date)
                    trip_df = pd.read_excel(files["trip"], dtype={"人员编号": str})

                    if files["shift"].name.endswith(".xlsx"):
                        shift_df = pd.read_excel(files["shift"], dtype={"工号": str})
                    else:
                        shift_df = pd.read_csv(files["shift"], encoding="gbk", dtype={"工号": str})

                    if files["record"].name.endswith(".csv"):
                        record_df = pd.read_csv(files["record"], encoding="gbk", parse_dates=["考勤时间"], dtype={"工号": str})
                    else:
                        record_df = pd.read_excel(files["record"], dtype={"工号": str})

                    # 处理PC考勤结果
                    date_range, attendance_data = process_pc_attendance(files["pc"])
                    contact_attendance_list, person_dept_dict = init_attendance_template(person_df, date_range[0], date_range[1])
                    index_map = build_record_index(contact_attendance_list)

                    fill_pc_attendance(index_map, attendance_data)
                    fill_oa_attendance(index_map, oa_df)
                    fill_leave_registration(index_map, leave_df)
                    fill_leave_info(index_map, qj_df)
                    fill_business_trip(index_map, trip_df)
                    shift_day_dict = fill_shift_attendance(index_map, shift_df, record_df, holiday_set, person_dept_dict)

                    # 汇总数据
                    summary_result = summarize_attendance(contact_attendance_list, holiday_set, shift_day_dict)
                    df_summary = pd.DataFrame(summary_result)
                    df_all = pd.DataFrame(contact_attendance_list)

                    # 清理数据
                    df_summary = clean_zeros(df_summary)
                    
                    # 保存带颜色标记的汇总表
                    save_excel_with_highlight(df_summary, "汇总表.xlsx")
                    
                    # 保存带颜色标记的明细表
                    save_excel_with_highlight(df_all, "明细表.xlsx")
                    
                    # 准备部门文件列表
                    dept_summary_files = []
                    dept_detail_files = []
                    
                    # 按一级部门分组并保存文件
                    if "部门" in df_summary.columns:
                        # 按一级部门分组
                        dept_groups_summary = df_summary.groupby(df_summary["部门"].astype(str).str.split("/").str[0])
                        dept_groups_detail = df_all.groupby(df_all["部门"].astype(str).str.split("/").str[0])
                        
                        # 为每个部门保存文件
                        for dept, group in dept_groups_summary:
                            dept_name = str(dept).strip().replace("/", "_").replace("\\", "_")
                            
                            # 保存部门汇总表
                            dept_summary_file = f"{dept_name}_汇总.xlsx"
                            save_excel_with_highlight(group, dept_summary_file)
                            dept_summary_files.append((dept_name, dept_summary_file))
                            
                            # 保存部门明细表
                            dept_detail_file = f"{dept_name}_明细.xlsx"
                            dept_detail_group = dept_groups_detail.get_group(dept)
                            save_excel_with_highlight(dept_detail_group, dept_detail_file)
                            dept_detail_files.append((dept_name, dept_detail_file))
                    
                    # 拆分原始打卡记录
                    split_files = split_attendance_records(files["record"], "原始打卡记录")
                    
                    # 创建ZIP文件
                    create_zip_file("考勤结果汇总.zip", "汇总表.xlsx", "明细表.xlsx", dept_summary_files, dept_detail_files, split_files)
                    
                    # 更新会话状态
                    st.session_state.analysis_completed = True
                    st.session_state.df_summary = df_summary
                    st.session_state.df_all = df_all
                    st.session_state.dept_summary_files = dept_summary_files
                    st.session_state.dept_detail_files = dept_detail_files
                    st.session_state.split_files = split_files
                    st.session_state.zip_file_created = True
                    
                    # 显示处理结果
                    st.success("✅ 考勤数据处理完成！")
                    st.write(f"📊 处理了 {len(contact_attendance_list)} 条考勤记录")
                    st.write(f"👥 涉及 {len(set(df_all['工号']))} 位员工")
                    st.write(f"⏱️ 用时 {time.time() - start_time:.2f} 秒")
                    
                except Exception as e:
                    st.error(f"❌ 处理过程中出现错误：{str(e)}")
                    st.exception(e)
        
        # 如果分析已完成，显示下载按钮
        if st.session_state.analysis_completed:
            # 提供下载链接
            st.subheader("💾 下载结果")
            
            # 提供下载整个结果的按钮
            if os.path.exists("考勤结果汇总.zip"):
                with open("考勤结果汇总.zip", "rb") as f:
                    st.download_button(
                        label="📥 下载整个结果（ZIP格式）",
                        data=f,
                        file_name="考勤结果汇总.zip",
                        mime="application/zip"
                    )
            
            # 提供下载汇总表和明细表的按钮
            if os.path.exists("汇总表.xlsx"):
                with open("汇总表.xlsx", "rb") as f:
                    st.download_button(
                        label="📥 下载汇总表",
                        data=f,
                        file_name="所有单位汇总表.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            if os.path.exists("明细表.xlsx"):
                with open("明细表.xlsx", "rb") as f:
                    st.download_button(
                        label="📥 下载明细表",
                        data=f,
                        file_name="所有单位明细表.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            # 清理临时文件按钮
            if st.button("🗑️ 清理临时文件", key="clean_temp_files"):
                # 清理部门文件
                for dept_name, file_path in st.session_state.dept_summary_files:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                
                for dept_name, file_path in st.session_state.dept_detail_files:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                
                # 清理拆分的原始打卡记录文件
                for file_path in st.session_state.get('split_files', []):
                    if os.path.exists(file_path):
                        os.remove(file_path)
                
                # 清理原始打卡记录目录
                if os.path.exists("原始打卡记录"):
                    shutil.rmtree("原始打卡记录")
                
                # 清理汇总表和明细表
                if os.path.exists("汇总表.xlsx"):
                    os.remove("汇总表.xlsx")
                
                if os.path.exists("明细表.xlsx"):
                    os.remove("明细表.xlsx")
                
                # 清理ZIP文件
                if os.path.exists("考勤结果汇总.zip"):
                    os.remove("考勤结果汇总.zip")
                
                # 重置会话状态
                st.session_state.analysis_completed = False
                st.session_state.df_summary = None
                st.session_state.df_all = None
                st.session_state.dept_summary_files = []
                st.session_state.dept_detail_files = []
                st.session_state.split_files = []
                st.session_state.zip_file_created = False
                
                st.success("✅ 临时文件已清理完成！")
                
                # 刷新页面
                st.rerun()

# 侧边栏信息
with st.sidebar:
    st.header("ℹ️ 关于")
    st.write("这是一个现代化的考勤分析工具，使用Streamlit框架构建。")
    st.write("支持批量上传考勤文件，并自动识别文件类型。")
    st.write("📅 更新日期：2025-01-01")
    st.write("🔧 版本：v2.0")
